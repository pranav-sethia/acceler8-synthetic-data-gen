import json
import openpyxl
import os

# --- CONFIGURATION: EDIT THESE VALUES FOR EACH NEW CAPABILITY ---
INPUT_JSON_FILE = "api_summary_results.json"
TEMPLATE_EXCEL_FILE = "template_SN.xlsx" 
OUTPUT_EXCEL_FILE = "Skilled_Navigator_Report.xlsx" 

# --- MAIN SCRIPT (No need to edit below this line) ---

def main():
    """
    Reads JSON results and populates a multi-sheet Excel report using a pre-formatted template,
    preserving the subcapability order from the template.
    """
    print(f"Reading results from '{INPUT_JSON_FILE}'...")
    try:
        with open(INPUT_JSON_FILE, 'r') as f:
            all_results = json.load(f)
    except FileNotFoundError:
        print(f"❌ Error: The file '{INPUT_JSON_FILE}' was not found.")
        return

    print(f"Loading template from '{TEMPLATE_EXCEL_FILE}'...")
    try:
        workbook = openpyxl.load_workbook(TEMPLATE_EXCEL_FILE)
        template_sheet = workbook.active
    except FileNotFoundError:
        print(f"❌ Error: Could not find the template file '{TEMPLATE_EXCEL_FILE}'.")
        return

    print(f"Found {len(all_results)} results. Populating Excel report...")
    
    # --- FIX: Read the subcapability order directly from the template ---
    template_subcap_order = []
    for row in template_sheet.iter_rows(min_row=5, max_col=3, values_only=True):
        subcap_name = row[2] # Column C is 'Sub Capabilities'
        if subcap_name and subcap_name not in template_subcap_order:
            template_subcap_order.append(subcap_name)
    print(f"Template subcapability order detected: {template_subcap_order}")
    
    original_template_name = template_sheet.title
    
    for result in all_results:
        persona_name = result['request_payload']['assessment_capability_results']['metadata']['employeeName']
        capability_data = result['request_payload']['assessment_capability_results']['capabilityData']
        
        summary_text = result['api_response'].get('summary', 'No summary generated.')
        summary_rules_list = result['api_response'].get('generation_details', {}).get('original_generation', {}).get('summary_rules', [])
        summary_rules = '\n'.join(summary_rules_list)

        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = persona_name
        
        current_row = 5
        is_first_row_for_capability = True
        
        # --- FIX: Loop using the order from the template, not alphabetical order ---
        for subcap_name in template_subcap_order:
            # Find the matching subcapability data from the JSON
            subcap_data = next((sc for sc in capability_data['subCapabilities'] if sc['name'] == subcap_name), None)
            if not subcap_data:
                continue

            is_first_row_for_subcapability = True
            for question in subcap_data['questions']:
                if question.get('employeeQuestion'):
                    new_sheet[f'A{current_row}'] = question['id']
                    if is_first_row_for_capability:
                        new_sheet[f'B{current_row}'] = capability_data['capability']
                        new_sheet[f'I{current_row}'] = capability_data['capabilityScores']['employeeScore']
                        new_sheet[f'J{current_row}'] = capability_data['capabilityScores']['employeeStage']
                    
                    if is_first_row_for_subcapability:
                        new_sheet[f'C{current_row}'] = subcap_name
                        new_sheet[f'K{current_row}'] = subcap_data.get('employeeScore', '')

                    new_sheet[f'D{current_row}'] = question.get('employeeQuestion', '')
                    new_sheet[f'G{current_row}'] = str(question.get('employeeResponse', '')).replace('\n', ' | ')
                    new_sheet[f'H{current_row}'] = question.get('employeeScore')
                    
                    is_first_row_for_capability = False
                    is_first_row_for_subcapability = False
                    current_row += 1

        new_sheet['T5'] = summary_text
        new_sheet['U5'] = summary_rules
        
        new_sheet['T5'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        new_sheet['U5'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

        print(f"  - Created and populated sheet for {persona_name}")

    if original_template_name in workbook.sheetnames:
        del workbook[original_template_name]

    workbook.save(OUTPUT_EXCEL_FILE)
    print(f"\n🚀 Success! The formatted report has been saved as '{OUTPUT_EXCEL_FILE}'")

if __name__ == '__main__':
    main()